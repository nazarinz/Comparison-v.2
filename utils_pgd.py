# utils_pgd.py — Shared helpers & constants for PGD Apps (full)
# -----------------------------------------------------------------------------
# Dipakai oleh:
#   - pages/PGD_Comparison.py
#   - pages/PO_Splitter.py
# Berisi:
#   • Konstanta styling Excel & preferensi kolom tanggal
#   • Helper I/O (read_excel_file, read_csv_file)
#   • Pipeline SAP/Infor: load, process, merge, clean, compare, export
#   • Utility Splitter: parse_input, normalize_items, chunk_list, dsb.
#   • Semua fungsi bebas-dependensi Streamlit agar mudah di-test.
# -----------------------------------------------------------------------------

from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ================== Warna, Kolom, Format ==================
INFOR_COLOR = "FFF9F16D"   # kuning lembut (header Infor)
RESULT_COLOR = "FFC6EFCE"  # hijau lembut (header Result_*)
OTHER_COLOR = "FFD9D9D9"   # abu-abu muda (header lainnya)
DATE_FMT = "m/d/yyyy"

INFOR_COLUMNS_FIXED = [
    "Order Status Infor", "Infor Quantity", "Infor Model Name", "Infor Article No",
    "Infor Classification Code", "Infor Delay/Early - Confirmation CRD",
    "Infor Delay - PO PSDD Update", "Infor Lead time", "Infor GPS Country",
    "Infor Ship-to Country", "Infor FPD", "Infor LPD", "Infor CRD", "Infor PSDD",
    "Infor PODD", "Infor PD",
]

DELAY_EMPTY_COLUMNS = [
    "Delay/Early - Confirmation CRD",
    "Infor Delay/Early - Confirmation CRD",
    "Result_Delay_CRD",
    "Delay - PO PSDD Update",
    "Infor Delay - PO PSDD Update",
]

DATE_COLUMNS_PREF = [
    "Document Date", "FPD", "LPD", "CRD", "PSDD", "FCR Date", "PODD", "PD", "PO Date", "Actual PGI",
    "Infor FPD", "Infor LPD", "Infor CRD", "Infor PSDD", "Infor PODD", "Infor PD",
]

# ================== Util: Waktu & I/O ==================
def today_str_id() -> str:
    """Tanggal hari ini zona Asia/Jakarta (UTC+7) dalam format YYYYMMDD."""
    return (datetime.utcnow() + timedelta(hours=7)).strftime("%Y%m%d")


def read_excel_file(file):
    """Baca Excel dengan engine openpyxl (kompatibel dengan Streamlit uploader)."""
    return pd.read_excel(file, engine="openpyxl")


def read_csv_file(file):
    """Baca CSV dengan fallback encoding umum."""
    for enc in ("utf-8", "utf-8-sig", "latin1"):
        try:
            file.seek(0)
            return pd.read_csv(file, encoding=enc)
        except Exception:
            continue
    file.seek(0)
    return pd.read_csv(file)


# ================== Util: Tanggal ==================
def convert_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Coerce kolom tanggal yang dikenal menjadi datetime (errors='coerce')."""
    for col in DATE_COLUMNS_PREF:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


# ================== SAP loaders ==================
def load_sap(sap_df: pd.DataFrame) -> pd.DataFrame:
    df = sap_df.copy()
    if "Quanity" in df.columns and "Quantity" not in df.columns:
        df.rename(columns={"Quanity": "Quantity"}, inplace=True)
    if "PO No.(Full)" in df.columns:
        df["PO No.(Full)"] = df["PO No.(Full)"].astype(str).str.strip()
    return convert_date_columns(df)


# ================== Infor loaders ==================
def load_infor_from_many_csv(csv_dfs, *, on_info=lambda msg: None, on_warn=lambda msg: None):
    """Gabungkan banyak CSV Infor; validasi kolom wajib; log via callback opsional."""
    data_list = []
    required_cols = [
        "PO Statistical Delivery Date (PSDD)",
        "Customer Request Date (CRD)",
        "Line Aggregator",
    ]
    for i, df in enumerate(csv_dfs, start=1):
        if all(col in df.columns for col in required_cols):
            data_list.append(df)
            on_info(f"Dibaca ✅ CSV ke-{i} (kolom wajib lengkap)")
        else:
            miss = [c for c in required_cols if c not in df.columns]
            on_warn(f"CSV ke-{i} dilewati ⚠️ (kolom wajib hilang: {miss})")
    if not data_list:
        return pd.DataFrame()
    return pd.concat(data_list, ignore_index=True)


# ================== Infor processing & Comparison ==================
def process_infor(df_all: pd.DataFrame) -> pd.DataFrame:
    """Ambil kolom penting dari Infor, agregasi per Order #, dan rename ke prefiks Infor.*"""
    selected_columns = [
        'Order #','Order Status','Model Name','Article Number','Gps Customer Number',
        'Country/Region','Customer Request Date (CRD)','Plan Date','PO Statistical Delivery Date (PSDD)',
        'First Production Date','Last Production Date','PODD','Production Lead Time',
        'Class Code','Delay - Confirmation','Delay - PO Del Update','Quantity'
    ]
    missing_cols = [col for col in selected_columns if col not in df_all.columns]
    if missing_cols:
        return pd.DataFrame()

    df_infor = df_all[selected_columns].copy()
    df_infor = df_infor.groupby('Order #', as_index=False).agg({
        'Order Status':'first','Model Name':'first','Article Number':'first','Gps Customer Number':'first',
        'Country/Region':'first','Customer Request Date (CRD)':'first','Plan Date':'first',
        'PO Statistical Delivery Date (PSDD)':'first','First Production Date':'first',
        'Last Production Date':'first','PODD':'first','Production Lead Time':'first',
        'Class Code':'first','Delay - Confirmation':'first','Delay - PO Del Update':'first',
        'Quantity':'sum'
    })
    df_infor["Order #"] = df_infor["Order #"].astype(str).zfill(10).str.strip()

    rename_cols = {
        'Order Status':'Order Status Infor','Model Name':'Infor Model Name','Article Number':'Infor Article No',
        'Gps Customer Number':'Infor GPS Country','Country/Region':'Infor Ship-to Country',
        'Customer Request Date (CRD)':'Infor CRD','Plan Date':'Infor PD',
        'PO Statistical Delivery Date (PSDD)':'Infor PSDD','First Production Date':'Infor FPD',
        'Last Production Date':'Infor LPD','PODD':'Infor PODD','Production Lead Time':'Infor Lead time',
        'Class Code':'Infor Classification Code','Delay - Confirmation':'Infor Delay/Early - Confirmation CRD',
        'Delay - PO Del Update':'Infor Delay - PO PSDD Update','Quantity':'Infor Quantity'
    }
    df_infor.rename(columns=rename_cols, inplace=True)
    return df_infor


def merge_sap_infor(df_sap: pd.DataFrame, df_infor: pd.DataFrame) -> pd.DataFrame:
    sap = df_sap.copy()
    inf = df_infor.copy()
    if 'PO No.(Full)' in sap.columns:
        sap['PO No.(Full)'] = sap['PO No.(Full)'].astype(str).str.zfill(10)
    if 'Order #' in inf.columns:
        inf['Order #'] = inf['Order #'].astype(str).str.zfill(10)
    return sap.merge(inf, how='left', left_on='PO No.(Full)', right_on='Order #')


def fill_missing_dates(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out['Order Status Infor'] = out.get('Order Status Infor', pd.Series(dtype=str)).astype(str).str.strip().str.upper()
    for col in ['LPD','FPD','CRD','PD','PSDD','PODD']:
        if col not in out.columns:
            out[col] = pd.NaT
        out[col] = pd.to_datetime(out[col], errors='coerce')
    mask_open = out['Order Status Infor'].eq('OPEN')
    min_dates = out[['CRD','PD']].min(axis=1)
    out.loc[mask_open & out['LPD'].isna(),'LPD'] = min_dates
    out.loc[mask_open & out['FPD'].isna(),'FPD'] = min_dates
    out.loc[mask_open & out['PSDD'].isna(),'PSDD'] = out['CRD']
    out.loc[mask_open & out['PODD'].isna(),'PODD'] = out['CRD']
    return out


def clean_and_compare(df_merged: pd.DataFrame) -> pd.DataFrame:
    df = df_merged.copy()

    # numerik
    for col in ["Quantity","Infor Quantity","Production Lead Time","Infor Lead time","Article Lead time"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).round(2)

    # mapping delay codes
    code_mapping = {
        '161':'01-0161','84':'03-0084','68':'02-0068','64':'04-0064','62':'02-0062','61':'01-0061',
        '51':'03-0051','46':'03-0046','7':'02-0007','3':'03-0003','2':'01-0002','1':'01-0001',
        '4':'04-0004','8':'02-0008','10':'04-0010','49':'03-0049','90':'04-0090','63':'03-0063'
    }
    def map_code_safely(x):
        try:
            return code_mapping.get(str(int(float(x))), x)
        except (ValueError, TypeError):
            return x

    if "Infor Delay/Early - Confirmation CRD" in df.columns:
        df["Infor Delay/Early - Confirmation CRD"] = (
            df["Infor Delay/Early - Confirmation CRD"].replace(['--','N/A','NULL'], pd.NA).apply(map_code_safely)
        )
    if "Infor Delay - PO PSDD Update" in df.columns:
        df["Infor Delay - PO PSDD Update"] = (
            df["Infor Delay - PO PSDD Update"].replace(['--','N/A','NULL'], pd.NA).apply(map_code_safely)
        )

    # normalisasi string
    string_cols = [
        "Model Name","Infor Model Name","Article No","Infor Article No",
        "Classification Code","Infor Classification Code",
        "Ship-to Country","Infor Ship-to Country",
        "Ship-to-Sort1","Infor GPS Country",
        "Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD",
        "Delay - PO PSDD Update","Infor Delay - PO PSDD Update"
    ]
    for col in string_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.upper()
    if "Ship-to-Sort1" in df.columns:
        df["Ship-to-Sort1"] = df["Ship-to-Sort1"].astype(str).str.replace(".0","", regex=False)
    if "Infor GPS Country" in df.columns:
        df["Infor GPS Country"] = df["Infor GPS Country"].astype(str).str.replace(".0","", regex=False)

    # hasil perbandingan
    def safe_result(c1, c2):
        if c1 in df.columns and c2 in df.columns:
            return np.where(df[c1] == df[c2], "TRUE", "FALSE")
        return ["COLUMN MISSING"] * len(df)

    df["Result_Quantity"]            = safe_result("Quantity","Infor Quantity")
    df["Result_Model Name"]          = safe_result("Model Name","Infor Model Name")
    df["Result_Article No"]          = safe_result("Article No","Infor Article No")
    df["Result_Classification Code"] = safe_result("Classification Code","Infor Classification Code")
    df["Result_Delay_CRD"]           = safe_result("Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD")
    df["Result_Delay_PSDD"]          = safe_result("Delay - PO PSDD Update","Infor Delay - PO PSDD Update")
    df["Result_Lead Time"]           = safe_result("Article Lead time","Infor Lead time")
    df["Result_Country"]             = safe_result("Ship-to Country","Infor Ship-to Country")
    df["Result_Sort1"]               = safe_result("Ship-to-Sort1","Infor GPS Country")
    df["Result_FPD"]                 = safe_result("FPD","Infor FPD")
    df["Result_LPD"]                 = safe_result("LPD","Infor LPD")
    df["Result_CRD"]                 = safe_result("CRD","Infor CRD")
    df["Result_PSDD"]                = safe_result("PSDD","Infor PSDD")
    df["Result_PODD"]                = safe_result("PODD","Infor PODD")
    df["Result_PD"]                  = safe_result("PD","Infor PD")

    return df


# ================== Kolom & Build Report ==================
DESIRED_ORDER = [
    'Client No','Site','Brand FTY Name','SO','Order Type','Order Type Description',
    'PO No.(Full)','Order Status Infor','PO No.(Short)','Merchandise Category 2','Quantity',
    'Infor Quantity','Result_Quantity','Model Name','Infor Model Name','Result_Model Name',
    'Article No','Infor Article No','Result_Article No','SAP Material','Pattern Code(Up.No.)',
    'Model No','Outsole Mold','Gender','Category 1','Category 2','Category 3','Unit Price',
    'Classification Code','Infor Classification Code','Result_Classification Code','DRC',
    'Delay/Early - Confirmation PD','Delay/Early - Confirmation CRD','Infor Delay/Early - Confirmation CRD',
    'Result_Delay_CRD','Delay - PO PSDD Update','Infor Delay - PO PSDD Update','Result_Delay_PSDD',
    'Delay - PO PD Update','MDP','PDP','SDP','Article Lead time','Infor Lead time',
    'Result_Lead Time','Cust Ord No','Ship-to-Sort1','Infor GPS Country','Result_Sort1',
    'Ship-to Country','Infor Ship-to Country','Result_Country',
    'Ship to Name','Document Date','FPD','Infor FPD','Result_FPD','LPD','Infor LPD',
    'Result_LPD','CRD','Infor CRD','Result_CRD','PSDD','Infor PSDD','Result_PSDD',
    'FCR Date','PODD','Infor PODD','Result_PODD','PD','Infor PD','Result_PD',
    'PO Date','Actual PGI','Segment','S&P LPD','Currency','Customer PO item'
]


def reorder_columns(df: pd.DataFrame, desired_order: list[str]) -> pd.DataFrame:
    existing = [c for c in desired_order if c in df.columns]
    tail = [c for c in df.columns if c not in existing]
    return df[existing + tail]


def build_report(df_sap: pd.DataFrame, df_infor_raw: pd.DataFrame) -> pd.DataFrame:
    df_infor = process_infor(df_infor_raw)
    if df_infor.empty:
        return pd.DataFrame()
    df_sap2 = convert_date_columns(load_sap(df_sap))
    df_infor2 = convert_date_columns(df_infor)
    df_merged = merge_sap_infor(df_sap2, df_infor2)
    df_merged = fill_missing_dates(df_merged)
    df_final = clean_and_compare(df_merged)
    return reorder_columns(df_final, DESIRED_ORDER)


# ================== Export Styled Excel ==================
def _blank_delay_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in DELAY_EMPTY_COLUMNS:
        if col in out.columns:
            out[col] = out[col].replace({
                np.nan:"", pd.NA:"", None:"", "NaN":"", "NAN":"",
                "NULL":"", "--":"", 0:"", 0.0:"", "0":""
            })
    return out


def _export_excel_styled(df: pd.DataFrame, sheet_name: str = "Report") -> io.BytesIO:
    """Header diwarnai; body plain; font Calibri 9; tanggal m/d/yyyy; auto width; freeze A2."""
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # bersihkan fill
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.fill = PatternFill(fill_type=None)

        # header coloring
        header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        idx_by_name = {c.value: i+1 for i, c in enumerate(header_cells)}
        for cell in header_cells:
            col_name = str(cell.value)
            if col_name in INFOR_COLUMNS_FIXED:
                fill = PatternFill("solid", fgColor=INFOR_COLOR)
            elif col_name.startswith("Result_"):
                fill = PatternFill("solid", fgColor=RESULT_COLOR)
            else:
                fill = PatternFill("solid", fgColor=OTHER_COLOR)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.font = Font(name="Calibri", size=9, bold=True)

        # body style
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.fill = PatternFill(fill_type=None)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
                cell.font = Font(name="Calibri", size=9)

        # format tanggal
        for date_col in DATE_COLUMNS_PREF:
            if date_col in idx_by_name:
                cidx = idx_by_name[date_col]
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=cidx)
                    if cell.value not in ("", None):
                        cell.number_format = DATE_FMT

        # auto width
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            maxlen = 0
            for cell in ws[col_letter]:
                v = "" if cell.value is None else str(cell.value)
                maxlen = max(maxlen, len(v))
            ws.column_dimensions[col_letter].width = min(max(9, maxlen + 2), 40)

        # UX
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    bio.seek(0)
    return bio


# ================== PO Splitter Helpers ==================
def parse_input(text: str, split_mode: str = "auto"):
    text = (text or "").strip()
    if not text:
        return []
    if split_mode == "newline":
        raw = text.splitlines()
    elif split_mode == "comma":
        raw = text.split(",")
    elif split_mode == "semicolon":
        raw = text.split(";")
    elif split_mode == "whitespace":
        raw = re.split(r"\s+", text)
    else:
        if "\n" in text:
            raw = re.split(r"[\r\n]+", text)
            split_more = []
            for line in raw:
                s = line.strip()
                if not s:
                    continue
                if ("," in s) or (";" in s):
                    split_more.extend(re.split(r"[,;]", s))
                else:
                    split_more.append(s)
            raw = split_more
        elif ("," in text) or (";" in text):
            raw = re.split(r"[,;]", text)
        else:
            raw = re.split(r"\s+", text)
    return [x.strip() for x in raw if str(x).strip() != ""]


def normalize_items(items, *, keep_only_digits=False, upper_case=False, strip_prefix_suffix=False):
    out = []
    for it in items:
        s = str(it)
        if strip_prefix_suffix:
            s = re.sub(r"^\W+|\W+$", "", s)
        if keep_only_digits:
            s = re.sub(r"\D+", "", s)
        if upper_case:
            s = s.upper()
        s = s.strip()
        if s:
            out.append(s)
    return out


def chunk_list(items, size):
    return [items[i:i+size] for i in range(0, len(items), size)]


def to_txt_bytes(lines):
    buf = io.StringIO()
    for ln in lines:
        buf.write(f"{ln}\n")
    return buf.getvalue().encode("utf-8")


def df_from_list(items, col_name="PO"):
    return pd.DataFrame({col_name: items})


def make_zip_bytes(chunks, *, basename="chunk", as_csv=True, col_name="PO"):
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for idx, part in enumerate(chunks, start=1):
            if as_csv:
                df = df_from_list(part, col_name=col_name)
                csv_bytes = df.to_csv(index=False).encode("utf-8")
                zf.writestr(f"{basename}_{idx:02d}.csv", csv_bytes)
            else:
                zf.writestr(f"{basename}_{idx:02d}.txt", to_txt_bytes(part))
    mem.seek(0)
    return mem
